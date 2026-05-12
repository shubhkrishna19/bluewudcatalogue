from __future__ import annotations

import argparse
import json
import re
import zipfile
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse
import xml.etree.ElementTree as ET

import openpyxl
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT.parent / "CHP Price list.xlsx"
DATA_DIR = ROOT / "data"
GENERATED_DIR = ROOT / "assets" / "generated"

MAIN_SHEET = "Price List Structured"
FIRST_PAGE_SHEET = "First Page "
LAST_PAGE_SHEET = "Last Page"
MASTER_SHEET_NAME = "MTP SKUs - Master Data"

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def column_name(cell_ref: str) -> str:
    return re.sub(r"\d+", "", cell_ref)


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"[ \t]+\n", "\n", text)


def clean_url(value) -> str:
    text = clean_text(value)
    return text.replace("\n", "").strip()


def normalize_tokens(value: str) -> set[str]:
    stopwords = {
        "and",
        "for",
        "with",
        "without",
        "the",
        "wood",
        "wooden",
        "engineered",
        "bluewud",
        "home",
        "furniture",
        "product",
        "products",
        "cabinet",
        "storage",
        "size",
    }
    tokens = set(re.findall(r"[a-z0-9]+", value.lower()))
    return {token for token in tokens if len(token) > 2 and token not in stopwords}


def strip_tracking_params(url: str) -> str:
    parsed = urlparse(url)
    kept_query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if not key.startswith("_") and key.lower() not in {"utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content"}
    ]
    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path.rstrip("/"),
            "",
            urlencode(kept_query),
            "",
        )
    )


def bluewud_search_url(product_name: str, mtp_sku: str) -> str:
    query = " ".join(part for part in [product_name, mtp_sku] if part)
    return f"https://www.bluewud.com/search?q={quote(query, safe='')}"


def resolve_product_url(raw_url: str, product_name: str, mtp_sku: str, sku_colors: str) -> dict[str, str]:
    fallback = bluewud_search_url(product_name, mtp_sku)
    url = clean_url(raw_url)
    if not url:
        return {
            "url": fallback,
            "sourceUrl": "",
            "status": "search_fallback_missing",
            "confidence": "low",
        }

    parsed = urlparse(url)
    hostname = parsed.netloc.lower().replace("www.", "")
    if hostname != "bluewud.com" or not parsed.path.startswith("/products/"):
        return {
            "url": fallback,
            "sourceUrl": url,
            "status": "search_fallback_external_or_non_product",
            "confidence": "low",
        }

    canonical = strip_tracking_params(urlunparse(("https", "www.bluewud.com", parsed.path, "", parsed.query, "")))
    handle = parsed.path.split("/products/", 1)[1].strip("/")
    source_tokens = normalize_tokens(" ".join([product_name, mtp_sku, sku_colors]))
    handle_tokens = normalize_tokens(handle.replace("-", " "))
    overlap = source_tokens & handle_tokens
    score = len(overlap) / max(1, min(len(source_tokens), len(handle_tokens)))

    # This guard prevents QR codes from sending customers to a clearly different
    # product when the workbook link has drifted. Exact live matching can be
    # added through the Bluewud product feed/Shopify Admin once credentials are available.
    if score < 0.2 and not any(token in handle for token in normalize_tokens(mtp_sku)):
        return {
            "url": fallback,
            "sourceUrl": canonical,
            "status": "search_fallback_suspect_excel_link",
            "confidence": "low",
        }

    return {
        "url": canonical,
        "sourceUrl": canonical,
        "status": "canonical_bluewud_product",
        "confidence": "medium",
    }


def format_number(value, decimals: int | None = None) -> str:
    if value in (None, ""):
        return ""
    try:
        number = Decimal(str(value))
    except Exception:
        return clean_text(value)

    if decimals is None:
        if number == number.to_integral():
            decimals = 0
        else:
            decimals = 2

    quant = Decimal("1") if decimals == 0 else Decimal("1." + ("0" * decimals))
    rounded = number.quantize(quant, rounding=ROUND_HALF_UP)
    return f"{rounded:,.{decimals}f}"


def format_excel_integer(value) -> str:
    if value in (None, ""):
        return ""
    try:
        rounded = Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        return f"{rounded:,}"
    except Exception:
        return clean_text(value)


def parse_external_master(workbook_path: Path) -> dict[str, dict[str, str]]:
    with zipfile.ZipFile(workbook_path) as zf:
        root = ET.fromstring(zf.read("xl/externalLinks/externalLink1.xml"))

    sheet_names = [node.attrib["val"] for node in root.findall(".//m:sheetName", NS)]
    try:
        master_sheet_id = str(sheet_names.index(MASTER_SHEET_NAME))
    except ValueError as exc:
        raise RuntimeError(f"Could not find external sheet {MASTER_SHEET_NAME!r}") from exc

    sheet_data = None
    for candidate in root.findall(".//m:sheetData", NS):
        if candidate.attrib.get("sheetId") == master_sheet_id:
            sheet_data = candidate
            break
    if sheet_data is None:
        raise RuntimeError(f"Could not find cached data for {MASTER_SHEET_NAME!r}")

    rows = sheet_data.findall("m:row", NS)
    if not rows:
        return {}

    def row_to_dict(row) -> dict[str, str]:
        result = {}
        for cell in row.findall("m:cell", NS):
            ref = cell.attrib.get("r", "")
            col = column_name(ref)
            result[col] = cell.findtext("m:v", default="", namespaces=NS)
        return result

    headers_by_col = {
        col: clean_text(value)
        for col, value in row_to_dict(rows[0]).items()
        if clean_text(value)
    }

    by_sku: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        raw = row_to_dict(row)
        item = {
            headers_by_col[col]: clean_text(value)
            for col, value in raw.items()
            if col in headers_by_col
        }
        sku = clean_text(item.get("MTP SKU"))
        if sku:
            by_sku[sku] = item

    return by_sku


def save_sheet_image(workbook, sheet_name: str, output_name: str) -> str:
    worksheet = workbook[sheet_name]
    images = getattr(worksheet, "_images", [])
    if not images:
        return ""

    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    output_path = GENERATED_DIR / output_name
    raw = images[0]._data()

    with Image.open(BytesIO(raw)) as img:
        img.save(output_path, format="PNG", optimize=True)

    return f"assets/generated/{output_path.name}"


def build_catalogue(workbook_path: Path) -> dict:
    master_by_sku = parse_external_master(workbook_path)

    value_wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    image_wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    ws = value_wb[MAIN_SHEET]

    cover_image = save_sheet_image(image_wb, FIRST_PAGE_SHEET, "front-page.png")
    back_image = save_sheet_image(image_wb, LAST_PAGE_SHEET, "last-page.png")
    logo_image = save_sheet_image(image_wb, MAIN_SHEET, "bluewud-logo.png")

    products = []
    for row_index in range(3, ws.max_row + 1):
        category = clean_text(ws.cell(row_index, 2).value)
        name = clean_text(ws.cell(row_index, 3).value)
        mtp_sku = clean_text(ws.cell(row_index, 11).value)
        if not any([category, name, mtp_sku]):
            continue

        master = master_by_sku.get(mtp_sku, {})
        website_url = clean_url(master.get("Website Link"))
        image_url = clean_url(master.get("Main Image Link"))
        sku_colors = clean_text(ws.cell(row_index, 4).value)
        resolved_link = resolve_product_url(website_url, name, mtp_sku, sku_colors)

        product = {
            "id": f"row-{row_index}",
            "rowNumber": row_index,
            "category": category,
            "itemName": name,
            "skuColors": sku_colors,
            "dimensionWeight": clean_text(ws.cell(row_index, 5).value),
            "weight": format_number(ws.cell(row_index, 6).value, 2),
            "mrp": format_excel_integer(ws.cell(row_index, 7).value),
            "chp": format_excel_integer(ws.cell(row_index, 8).value),
            "mtpSku": mtp_sku,
            "sourceWebsiteUrl": website_url,
            "websiteUrl": resolved_link["url"],
            "linkStatus": resolved_link["status"],
            "linkConfidence": resolved_link["confidence"],
            "imageUrl": image_url,
            "qrUrl": f"https://quickchart.io/qr?text={quote(resolved_link['url'], safe='')}" if resolved_link["url"] else "",
            "searchText": " ".join(
                part
                for part in [
                    category,
                    name,
                    clean_text(ws.cell(row_index, 4).value),
                    clean_text(ws.cell(row_index, 5).value),
                    mtp_sku,
                    website_url,
                ]
                if part
            ).lower(),
        }
        products.append(product)

    categories: dict[str, dict[str, object]] = {}
    for product in products:
        key = product["category"].strip().lower()
        if not key:
            continue
        if key not in categories:
            categories[key] = {"label": product["category"], "count": 0}
        categories[key]["count"] = int(categories[key]["count"]) + 1

    source_stat = workbook_path.stat()
    return {
        "meta": {
            "title": "CHP Price Catalogue",
            "brand": "Bluewud",
            "sourceFile": workbook_path.name,
            "sourceLastModified": datetime.fromtimestamp(source_stat.st_mtime, timezone.utc).isoformat(),
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "productCount": len(products),
            "sheetNames": value_wb.sheetnames,
            "coverImage": cover_image,
            "backImage": back_image,
            "logoImage": logo_image,
            "sharePointSource": "https://bluewud.sharepoint.com/:x:/r/sites/GraphicUnit/_layouts/15/Doc.aspx?sourcedoc=%7B6E7AF49D-2928-4F00-94BD-FE9CE82FA07F%7D&file=CHP%20Price%20list.xlsx&fromShare=true&action=default&mobileredirect=true",
        },
        "categories": sorted(categories.values(), key=lambda item: str(item["label"]).lower()),
        "products": products,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract the CHP workbook into static catalogue data.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)

    catalogue = build_catalogue(workbook_path)
    output_path = DATA_DIR / "catalogue.json"
    output_path.write_text(json.dumps(catalogue, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Wrote {output_path}")
    print(f"Products: {catalogue['meta']['productCount']}")
    print(f"Cover: {catalogue['meta']['coverImage']}")
    print(f"Back: {catalogue['meta']['backImage']}")


if __name__ == "__main__":
    main()
